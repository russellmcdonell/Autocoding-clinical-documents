# pylint: disable=line-too-long, broad-exception-caught, too-many-lines, too-many-nested-blocks, wrong-import-position
'''
Script AutoCoding.py
A script to find clinical concepts (MetaThesaurus codes) in text document and map them to other coding systems.

This script import modules for the relevant text document type; a module for preparing the document text for autocoding,
a module for completing any coding not completed by MetaMapLite (extending negation etc),
and a module for analysing the required codes.

This script reads a text document from <stdin> or reads a text documents from a file(s),
or reads text documents from a folder and autocodes it/them and prints the extracted codes.

Labels and lists are processed before Terms.
History section and history markers are found before concepts are extracted.
Modifiers are done


    SYNOPSIS
    $ python AutoCode.py [-I inputDir|--inputDir=inputDir]
        [-i inputFile|--inputFile=inputFile]
        [-O outputDir|--outputDir=outputDir]
        [-C configDir|--configDir=configDir]
        [-c configFile|--configFile=configFile]
        [-v loggingLevel|--verbose=logingLevel]
        [-L logDir|--logDir=logDir]
        [-l logfile|--logfile=logfile]
        [-|filename]...


    REQUIRED


    OPTIONS
    -I inputDir|--inputDir=inputDir
    The folder containing the text document(s).

    -I inputFile|--inputFile=inputFile
    The name of the clinical document to be AutoCoded.

    -O outputDir|--outputDir=outputDir
    The foleder where the output file(s) will be created.

    -H hostname|--MetaMapLiteHost=hostname
    The name of the MetaMapLite Server (default="localhost")

    -P port|--MetaMapLitePort=port
    The port for the AutoCoding service on the MetaMapLite Server (default="8080")

    -U URL|--MetaMapLiteURL=URL
    The URL for the AutoCoding service on the MetaMapLite Server (default="/AutoCoding/MetaMapLite")

    -v loggingLevel|--verbose=loggingLevel
    Set the level of logging that you want.

    -L logDir|--logDir=logDir
    The directory where the log file will be created (default=".").

    -l logfile|--logfile=logfile
    The name of a log file where you want all messages captured.


    This script processes a clinical document, which may contain headings, lists, wrapped text etc.
    This script uses MetaMapLite to identify the clinical concepts within the clinical document and
    also split the clinical document into sentences (full stops are good!).

    However, before this the clinical document may need cleaning (preparing) so that MetaMapLite will be able
    to find all the clinical terms. For instance, abbrevations and acronyms will have to be replaced with their
    full, correct medical terms.

    And after coding, all we will have it a collection of clinical codes. Some processing will be required in order
    to create a collection of interpretable/analyzable clinical concepts. Firstly, it may be necessary to
    interpret the codes "in context" so we create a new "document" being a blend of sentences and
    clinical codes embedded within those sentences at the point within each sentence where MetaMapLite
    identified as piece of text that mapped to a clinical code. Whilst doing this we identify which
    sentences or parts of sentences contain historical information ('past history of ...'). Usually, any analysis
    will want to focus on current conditions/problems and ignore any historical observations.

    Ths other "in context" issue that will need addressing is 'negation'. MetaMapLite will identify negated concepts
    such as "no signs of X" but only negated concepts. MetaMapLite does not propate negation to subsequent concepts.
    The AutoCoding Clinical Documents project has to be able to handle sentences like "no signs of X or Y,
    but clear evidence of Z". The negation of "X" must be extended to "Y", but negation extension must cease
    whenever the word "but" or "however" or any other, similar negation termination word or phrase is encountered.

    The final step in creating a set of interpretable/analyzable clinical concepts it to recognize that MetaMapLite
    returns codes with varing degrees of specificity; sometimes it returns a sequence of small concepts which,
    when taken as a whole, imply a more complex concept that is relevant to any subsequent analysis.

    The AutoCoding Clinical Documents project handles all of this at an abtract level, which means that it can
    be used in many, varied situations, to automatically code many, varied types of clinical documents. However
    the details will vary greatly between Use Cases; the acronyms found in nursing notes will be different to
    those found in discharge sumarries which, in turn, will be different to those found in pathology reports.
    To cater for these different Use Cases, the AutoCoding Clinical Documents project read solution specific
    configuration data and imports solution specific Python modules, for each Use Case. The name of the "solution"
    (name of the folder containing the solution specific Excel workbooks and Python scripts) must be specified
    on the command line when you run "AutoCoding.py".    
'''

# pylint: disable=invalid-name, bare-except, pointless-string-statement, unspecified-encoding

import os
import sys
import logging
import importlib
import threading
import re
from http import client
from openpyxl import load_workbook
from flask import Flask
import functions as f
import excelFunctions as excel
import data as d


app = Flask(__name__)
import flaskFunctions as flsk


def getDocument(fileName):
    '''
    Get a clinical from a file or standard input
    '''
    d.rawClinicalDocument = ''
    if fileName == '-':     # Use standard input
        for line in sys.stdin:
            d.rawClinicalDocument += line.rstrip() + '\n'
        return
    with open(fileName, 'rt', newline='', encoding='utf-8') as fp:
        for line in fp:
            d.rawClinicalDocument += line.rstrip() + '\n'
        return



if __name__ == '__main__':
    '''
    The main code
    Start by parsing the command line arguements and setting up logging.
    Then import the solution files and read in the configuration data.
    Validate the solution - all files exist, all worksheets in workbooks.
    Then process each file name in the command line - read the text document,
    autocode it, analyze the sentences and codes before printing the results.
    '''

    # Quick scan of the command line options to see if '-F' or '--Flask' have been set
    isFlask = False
    for arg in sys.argv:
        if arg == '-F':
            isFlask = True
        elif arg == '--Flask':
            isFlask = True

    # Set the command line options
    parser = f.setArguments(isFlask)

    # Parse the command line
    f.doArguments(parser, isFlask)

    # Check for a valid combination of arguments
    if (d.inputFile is not None) and (d.inputDir is None):
        logging.critical('Must specify inputDir when specifying inputFile')
        logging.shutdown()
        sys.exit(d.EX_CONFIG)

    # Make sure we can connect to the MetaMapLite Service
    MetaMapLiteConnection = None                # The MetaMapLite http connection
    d.MetaMapLiteServiceLock = threading.Lock()    # A Threading Lock used to avoid threading issues in the MetaMapLiteService
    d.MetaMapLiteHeaders = {'Content-type':'application/x-www-form-urlencoded', 'Accept':'application/json'}
    try:
        MetaMapLiteConnection = client.HTTPConnection(d.MetaMapLiteHost, d.MetaMapLitePort)
        MetaMapLiteConnection.close()
    except (client.NotConnected, client.InvalidURL, client.UnknownProtocol,client.UnknownTransferEncoding,client.UnimplementedFileMode,
            client.IncompleteRead, client.ImproperConnectionState, client.CannotSendRequest, client.CannotSendHeader,
            client.ResponseNotReady, client.BadStatusLine) as e:
        logging.critical('Cannot connect to the MetaMapLite Service on host (%s) and port (%s) [error:(%s)]',
                         d.MetaMapLiteHost, d.MetaMapLitePort, repr(e))
        logging.shutdown()
        sys.stdout.flush()
        sys.exit(d.EX_UNAVAILABLE)
    logging.debug('Connection to MetaMapLite server tested')

    # Check that the solution files all exist and appear correct.
    if not os.path.isdir(os.path.join('solutions', d.solution)):
        logging.critical('No solution folder named "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    if not os.path.isfile(os.path.join('solutions', d.solution, 'solutionData.py')):
        logging.critical('No solution file "solutionData.py" in solution folder "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    if not os.path.isfile(os.path.join('solutions', d.solution, 'prepare.py')):
        logging.critical('No solution file "prepare.py" in solution folder "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    if not os.path.isfile(os.path.join('solutions', d.solution, 'prepare.xlsx')):
        logging.critical('No solution file "prepare.xlsx" in solution folder "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    if not os.path.isfile(os.path.join('solutions', d.solution, 'complete.py')):
        logging.critical('No solution file "complete.py" in solution folder "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    if not os.path.isfile(os.path.join('solutions', d.solution, 'complete.xlsx')):
        logging.critical('No solution file "complete.xlsx" in solution folder "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    if not os.path.isfile(os.path.join('solutions', d.solution, 'analyze.py')):
        logging.critical('No solution file "analyze.py" in solution folder "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    if not os.path.isfile(os.path.join('solutions', d.solution, 'analyze.xlsx')):
        logging.critical('No solution file "analyze.xlsx" in solution folder "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    if not os.path.isfile(os.path.join('solutions', d.solution, 'Solution MetaThesaurus.xlsx')):
        logging.critical('No solution file "Solution MetaThesaurus.xlsx" in solution folder "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)

    # Check the solution specific MetaThesaurus Excel workbook
    wb = load_workbook(os.path.join('solutions', d.solution, 'Solution MetaThesaurus.xlsx'))
    requiredColumns = ['MetaThesaurus code', 'MetaThesaurus description','Source', 'Source code']
    this_df = excel.checkWorksheet(wb, 'Solution MetaThesaurus', 'Solution MetaThesaurus', requiredColumns, True)
    this_df = this_df.set_index(this_df.iloc[:,0].name)     # The code is really the index/dictionary keys()
    this_df = this_df.rename(columns={'MetaThesaurus_description': 'description'})
    d.solutionMetaThesaurus = this_df.to_dict(orient='index')

    # Import the solution specific data module
    try:
        d.sd = importlib.import_module('solutions.' + d.solution + '.solutionData')
    except Exception as e:
        logging.fatal('Cannot import "solutionData.py" for solution "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    logging.debug('Solution specific data loaded')

    # Check the solution 'prepare' Excel workbook
    wb = load_workbook(os.path.join('solutions', d.solution, 'prepare.xlsx'))
    requiredColumns = ['Label', 'Replacement']
    d.labels = excel.loadSimpleCompileSheet(wb, 'prepare', 'labels', requiredColumns, r'^', None, False, False)
    requiredColumns = ['Common', 'Technical']
    d.terms = excel.loadSimpleCompileSheet(wb, 'prepare', 'terms', requiredColumns, None, None, True, True)
    requiredColumns = ['Preamble markers', 'isCase', 'isStart']
    d.preambleMarkers = excel.loadBoolCompileWorksheet(wb, 'prepare', 'preamble markers', requiredColumns, None, None, True)
    requiredColumns = ['Preamble terms', 'Technical']
    d.preambleTerms = excel.loadSimpleCompileSheet(wb, 'prepare', 'preamble terms', requiredColumns, None, None, True, True)

    # Import the solution specific prepare module and check that it has all the required functions
    try:
        d.sp = importlib.import_module('solutions.' + d.solution + '.prepare')
    except Exception as e:
        logging.fatal('Cannot import "prepare.py" for solution "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    for requiredFunction in ['configure', 'solutionCleanDocument', 'solutionCheckPreamble', 'solutionCheckNotPreamble']:
        if not hasattr(d.sp, requiredFunction):
            logging.fatal('"prepare" module in solution "%s" is missing the "%s" function', d.solution, requiredFunction)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)

    # Now do any solution specific 'prepare' configuration and initialzation
    configConcepts = d.sp.configure(wb)
    d.knownConcepts.update(configConcepts)
    logging.debug('Prepare module loaded and configured')

    # Check the solution 'complete' Excel workbook
    wb = load_workbook(os.path.join('solutions', d.solution, 'complete.xlsx'))
    requiredColumns = ['History markers', 'isCase', 'isStart']
    d.historyMarkers = excel.loadBoolCompileWorksheet(wb, 'complete', 'history markers', requiredColumns, None, None, True)
    requiredColumns = ['Pre history']
    d.preHistory = excel.loadSimpleCompileSheet(wb, 'complete', 'pre history', requiredColumns, None, None, True, True)
    requiredColumns = ['Section markers', 'Section', 'isCase']
    d.sectionMarkers = excel.loadSimpleCompileSheet(wb, 'complete', 'section markers', requiredColumns, None, None, True, True)
    requiredColumns = ['SolutionID', 'MetaThesaurusID(s)']
    this_df = excel.checkWorksheet(wb, 'complete', 'equivalents', requiredColumns, False)
    thisData = this_df.values.tolist()
    for record in thisData:
        if record[0] is None:
            break
        # logging.debug("sheet(site implied), columns(%s), record(%s)", requiredColumns, record)
        equivalent = record[0]
        d.knownConcepts.add(equivalent)
        j = 1
        while (j < len(record)) and (record[j] is not None):
            d.equivalents[record[j]] = equivalent
            d.knownConcepts.add(record[j])
            j += 1
    requiredColumns = ['But boundaries']
    d.butBoundaries = excel.loadSimpleCompileSheet(wb, 'complete', 'but boundaries', requiredColumns, None, None, True, True)
    requiredColumns = ['Pre negations']
    d.preNegation = excel.loadCompileConceptsWorksheet(wb, 'complete', 'pre negations', requiredColumns, None, r'.*')
    requiredColumns = ['Immediate pre negations']
    d.immediatePreNegation = excel.loadCompileConceptsWorksheet(wb, 'complete', 'immediate pre negations', requiredColumns, None, r'\s+')
    requiredColumns = ['Post negations', 'Exceptions']
    d.postNegation = excel.loadCompileCompileWorksheet(wb, 'complete', 'post negations', requiredColumns, r'.*')
    requiredColumns = ['Immediate post negations', 'Exceptions']
    d.immediatePostNegation = excel.loadCompileCompileWorksheet(wb, 'complete', 'immediate post negations', requiredColumns, r'\s*')
    requiredColumns = ['Pre ambiguous']
    d.preAmbiguous = excel.loadCompileConceptsWorksheet(wb, 'complete', 'pre ambiguous', requiredColumns, None, r'.*')
    requiredColumns = ['Immediate pre ambiguous']
    d.immediatePreAmbiguous = excel.loadCompileConceptsWorksheet(wb, 'complete', 'immediate pre ambiguous', requiredColumns, None, r'\s+')
    requiredColumns = ['Post ambiguous', 'Exceptions']
    d.postAmbiguous = excel.loadCompileCompileWorksheet(wb, 'complete', 'post ambiguous', requiredColumns, r'\s*')
    requiredColumns = ['Immediate post ambiguous', 'Exceptions']
    d.immediatePostAmbiguous = excel.loadCompileCompileWorksheet(wb, 'complete', 'immediate post ambiguous', requiredColumns, r'\s*')
    requiredColumns = ['MetaThesaurusID', 'SolutionID', 'Modifier']
    this_df = excel.checkWorksheet(wb, 'complete', 'pre modifiers', requiredColumns, True)
    d.preModifiers = excel.loadModifierWorksheet(wb, 'complete', 'pre modifiers', requiredColumns, None, r'\s.*$')
    d.postModifiers = excel.loadModifierWorksheet(wb, 'complete', 'post modifiers', requiredColumns, r'^\s*', None)
    requiredColumns = ['SolutionID', 'Concept']
    this_df = excel.checkWorksheet(wb, 'complete', 'sentence concepts', requiredColumns, True)
    for row in this_df.itertuples():
        if row.SolutionID is None:
            break
        # logging.debug("sheet(sentence concepts)), columns(%s), row(%s)", requiredColumns, row)
        concept, isNeg = excel.checkConfigConcept(row.SolutionID)
        reText = re.compile(excel.checkPattern(row.Concept), flags=re.IGNORECASE|re.DOTALL)
        text = row.Concept
        text = re.sub(r'\s+', ' ', text)
        text = re.sub(r'\b', '', text)
        text = re.sub(r'\(', '(', text)
        text = re.sub(r'\)', ')', text)
        d.sentenceConcepts.append((concept, reText, isNeg, text))
        d.knownConcepts.add(concept)
    requiredColumns = ['Start Negation', 'End Negation', 'Sentences']
    this_df = excel.checkWorksheet(wb, 'complete', 'gross negations', requiredColumns, True)
    for row in this_df.itertuples():
        if row.Start_Negation is None:
            break
        # logging.debug("sheet(gross negations)), columns(%s), row(%s)", requiredColumns, row)
        startNeg = re.compile(excel.checkPattern(row.Start_Negation), flags=re.IGNORECASE|re.DOTALL)
        endNeg = re.compile(excel.checkPattern(row.End_Negation), flags=re.IGNORECASE|re.DOTALL)
        sentences = int(row.Sentences)
        d.grossNegation.append((startNeg, endNeg, sentences))
    requiredColumns = ['SolutionID', 'Section', 'Negate', 'MetaThesaurusIDs']
    d.sentenceNegationLists = excel.loadNegationListWorksheet(wb, 'complete', 'sentence negation lists', requiredColumns)
    d.documentNegationLists = excel.loadNegationListWorksheet(wb, 'complete', 'document negation lists', requiredColumns)
    requiredColumns = ['SolutionID', 'Asserted', 'MetaThesaurus or Solution IDs']
    d.sentenceConceptSequenceSets = excel.loadSequenceConceptSetsWorksheet(wb, 'complete', 'sent strict seq concept sets', requiredColumns, True)
    d.sentenceConceptSequenceSets += excel.loadSequenceConceptSetsWorksheet(wb, 'complete', 'sentence sequence concept sets', requiredColumns, False)
    requiredColumns = ['SolutionID', 'Sentences', 'Asserted', 'MetaThesaurus or Solution IDs']
    d.sentenceConceptSequenceSets += excel.loadSequenceConceptSetsWorksheet(wb, 'complete', 'multi sent strict seq conc sets', requiredColumns, True)
    d.sentenceConceptSequenceSets += excel.loadSequenceConceptSetsWorksheet(wb, 'complete', 'multi sentence seq concept sets', requiredColumns, False)
    requiredColumns = ['SolutionID', 'Asserted', 'MetaThesaurus or Solution IDs']
    d.sentenceConceptSets = excel.loadConceptSetsWorksheet(wb, 'complete', 'sentence concept sets', requiredColumns)
    requiredColumns = ['SolutionID', 'Sentences', 'Asserted', 'MetaThesaurus or Solution IDs']
    d.sentenceConceptSets += excel.loadConceptSetsWorksheet(wb, 'complete', 'multi sentence concept sets', requiredColumns)
    requiredColumns = ['SolutionID', 'Asserted', 'MetaThesaurus or Solution IDs']
    d.documentConceptSequenceSets = excel.loadConceptSetsWorksheet(wb, 'complete', 'document sequence concept sets', requiredColumns)
    d.documentConceptSets = excel.loadConceptSetsWorksheet(wb, 'complete', 'document concept sets', requiredColumns)
    requiredColumns = ['MetaThesaurus code']
    this_df = excel.checkWorksheet(wb, 'complete', 'other concepts', requiredColumns, True)
    d.otherConcepts = set(this_df['MetaThesaurus_code'].unique())


    # Import the solution specific complete module and check that it has all the required functions
    try:
        d.sc = importlib.import_module('solutions.' + d.solution + '.complete')
    except Exception as e:
        logging.fatal('Cannot import "complete.py" for solution "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    for requiredFunction in ['configure', 'requireConcept', 'solutionCheckHistory', 'addRawConcepts', 'initalizeNegation',
                             'extendNegation', 'higherConceptFound', 'setConcept', 'solutionAddAdditionalConcept', 'addFinalConcepts', 'complete']:
        if not hasattr(d.sc, requiredFunction):
            logging.fatal('"complete" module in solution "%s" is missing the "%s" function', d.solution, requiredFunction)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)

    # Now do any solution specific 'complete' configuration and initialzation
    configConcepts = d.sc.configure(wb)
    d.knownConcepts.update(configConcepts)
    logging.debug('Complete module loaded and configured')

    # Check the solution 'analyze' Excel workbook
    wb = load_workbook(os.path.join('solutions', d.solution, 'analyze.xlsx'))

    # There are no standard solution 'analyze' Excel worksheets

    # Import the solution specific analyze module and check that it has all the required functions
    try:
        d.sa = importlib.import_module('solutions.' + d.solution + '.analyze')
    except Exception as e:
        logging.fatal('Cannot import "analyze.py" for solution "%s"', d.solution)
        logging.shutdown()
        sys.exit(d.EX_CONFIG)
    for requiredFunction in ['configure', 'analyze']:
        if not hasattr(d.sa, requiredFunction):
            logging.fatal('"analyzee" module in solution "%s" is missing the "%s" function', d.solution, requiredFunction)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)
    if isFlask:
        for requiredFunction in ['reportHTML', 'reportJSON']:
            if not hasattr(d.sa, requiredFunction):
                logging.fatal('"analyze" module in solution "%s" is missing the "%s" function', d.solution, requiredFunction)
                logging.shutdown()
                sys.exit(d.EX_CONFIG)
    else:
        if not hasattr(d.sa, 'reportFile'):
            logging.fatal('"analyzee" module in solution "%s" is missing the "reportFile" function', d.solution)
            logging.shutdown()
            sys.exit(d.EX_CONFIG)

    # Now do any solution specific 'analyze' configuration and initialzation
    configConcepts = d.sa.configure(wb)
    d.knownConcepts.update(configConcepts)
    logging.debug('Analyze module loaded and configured')

    if isFlask:         # Run as a website and api service
        print(f'flsk:{flsk}')
        app.run(host="0.0.0.0")
        sys.exit(d.EX_OK)

    # Not flask, so read in the file(s) and AutoCode them
    # If neither inputDir, nor inputFile is specified, then read one file from standard input
    # and print the results to standard output.
    # If only inputDir is specified and outputDir is specified then process all the files in inputDir
    # and print all the results to outputDir with the same filename.
    # Otherwise, process just the one file - inputDir/inputFile. If outputDir is specified
    # then print the result to outputDir/inputFile, otherwise print the result to inputDir/AutoCoded_inputFile
    files = []
    if d.inputDir is None:
        files.append('-')
    elif d.inputFile is not None:
        files.append(d.inputFile)
    else:
        files = os.listdir(d.inputDir)

    for file in files:
        if d.inputDir is None:
            getDocument(file)
        else:
            getDocument(os.path.join(d.inputDir, file))

        # AutoCode this clinical document
        success = f.AutoCode()
        if success != d.EX_OK:
            continue

        # Print this clinical document
        if file == '-':
            d.sa.reportFile(None, None)
            sys.exit(d.EX_OK)
        baseFile = os.path.basename(file)
        filePart, ext = os.path.splitext(baseFile)
        if d.outputDir is None:
            d.sa.reportFile(d.inputDir, 'AutoCoded_' + filePart)
        else:
            d.sa.reportFile(d.outputDir, filePart)
